from selenium import webdriver
import string
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.ie.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import time 

import openpyxl
t ='';
vk = openpyxl.load_workbook(r'C:\Users\Chetan.K\Desktop\PTC_activity\Safa_Face_2\Code\Running_files\Safa_Primary1.xlsx')
sh = vk.active
m = sh.max_row
print(m)
# Hard coding values
#Country_hardcore = 'Germany'
Address_line1_hardcore = 'Personal client info - EY team only'
Address_line2_hardcore = ''
City_hardcore = 'Unknown'
Postal_Code_hardcore = '1111'
PO_Box_hardcore = ''

for i in range(2, m+1):
    try:
        cap = DesiredCapabilities().INTERNETEXPLORER
        cap['ignoreProtectedModeSettings'] = True
        cap['IntroduceInstabilityByIgnoringProtectedModeSettings'] = True
        cap['nativeEvents'] = True
        cap['ignoreZoomSetting'] = True
        cap['requireWindowFocus'] = True
        cap['INTRODUCE_FLAKINESS_BY_IGNORING_SECURITY_DOMAINS'] = True
        driver = webdriver.Ie(capabilities=cap, executable_path = r"C:\Users\Chetan.K\PycharmProjects\Office_Project\IEDriverServer.exe")
        
        value = sh.cell(i, 1).value
        driver.get(f'https://mdm.ey.net/businesspartnermaintenance/change/address/{value}') 
            
        try:
            driver.implicitly_wait(20) 
            driver.find_element_by_xpath("//*[@id='ddl-change-type-address']/option[3]").click()
            val1 = driver.find_element_by_xpath("//*[@id='ddl-existing-address']")
            for option in val1.find_elements_by_tag_name('option'):
                if 'P [Code:' in option.text:
                    t = option.text
                if option.text == t:
                    option.click()
                    
                    #Fething the values from primary address
                    address_line1 = driver.find_element_by_xpath("//*[@id='Address_AddressLine1']").get_attribute('value')
                    address_line2 = driver.find_element_by_xpath("//*[@id='Address_AddressLine2']").get_attribute('value')
                    address_line3 = driver.find_element_by_xpath("//*[@id='Address_AddressLine3']").get_attribute('value')
                    country = driver.find_element_by_id('ddl-country').get_attribute('value')
                    State = driver.find_element_by_id('ddl-state').get_attribute('value')
                    city = driver.find_element_by_xpath("//*[@id='Address_City']").get_attribute('value')
                    Postal_code = driver.find_element_by_xpath("//*[@id='Address_PostalCode']").get_attribute('value')
                    PO_box = driver.find_element_by_xpath("//*[@id='Address_PoBox']").get_attribute('value')
                    
                    #Primary values storing into Excel sheet
                    sh.cell(i, 5).value = address_line1
                    sh.cell(i, 6).value = address_line2
                    sh.cell(i, 7).value = address_line3
                    sh.cell(i, 8).value = country
                    sh.cell(i, 9).value = State
                    sh.cell(i, 10).value = city
                    sh.cell(i, 11).value = Postal_code
                    sh.cell(i, 12).value = PO_box
                   
                    # Clearing values of Primary address and inserting Masking Address 
                    if country in ['CH','AT','DE']:
                        driver.find_element_by_xpath("//*[@id='Address_AddressLine1']").clear()
                        driver.find_element_by_xpath("//*[@id='Address_AddressLine1']").send_keys(Address_line1_hardcore)
                        driver.find_element_by_xpath("//*[@id='Address_AddressLine2']").clear()
                        driver.find_element_by_xpath("//*[@id='Address_AddressLine3']").clear()
#                        driver.find_element_by_xpath("//*[@id='Address_AddressLine4']").clear()
#                        driver.find_element_by_xpath("//*[@id='Address_AddressLine5']").clear()
                        time.sleep(3)
                        driver.find_element_by_id('ddl-country').send_keys('USA')
                        time.sleep(2)                        
                        if country == 'CH':
                            driver.find_element_by_id('ddl-country').send_keys('Switzerland')
                        elif country == 'AT':
                            driver.find_element_by_id('ddl-country').send_keys('Austria')
                        else:
                            driver.find_element_by_id('ddl-country').send_keys('Germany')
                        #driver.find_element_by_xpath("//*[@id='ddl-state']").send_keys(State)
                        driver.find_element_by_xpath("//*[@id='Address_City']").clear()
                        driver.find_element_by_xpath("//*[@id='Address_City']").send_keys(City_hardcore)
                        driver.find_element_by_xpath("//*[@id='Address_PostalCode']").clear()
                        if country == 'DE':
                            driver.find_element_by_xpath("//*[@id='Address_PostalCode']").send_keys('11111')
                        else:
                            driver.find_element_by_xpath("//*[@id='Address_PostalCode']").send_keys(Postal_Code_hardcore)
                        driver.find_element_by_xpath("//*[@id='Address_PoBox']").clear()
                        driver.find_element_by_id('btn-submit').send_keys(Keys.ENTER)
                        time.sleep(2)
                        print(i-1)
                        sh.cell(i, 4).value = "updated as primary"
                        sh.cell(i, 2).value = "successful"
                        vk.save(r'C:\Users\Chetan.K\Desktop\PTC_activity\Safa_Face_2\Code\Running_files\Safa_Primary1.xlsx')
                    else:
                        print(i-1)
                        sh.cell(i, 4).value = "Non CH or AT country"
                        sh.cell(i, 2).value = "unsuccessful"
                        vk.save(r'C:\Users\Chetan.K\Desktop\PTC_activity\Safa_Face_2\Code\Running_files\Safa_Primary1.xlsx')
                        
        except Exception as e:
            #print('Some Issue with ' + sh.cell(i, 1).value)
            #sh.cell(i, 2).value = "Unsuccessful"
            vk.save(r"C:\Users\Chetan.K\Desktop\PTC_activity\Safa_Face_2\Code\Running_files\Safa_Primary1.xlsx")
        finally:
            driver.close()            
                    
    except Exception as e:
        print('Some Issue with internet')
        sh.cell(i, 2).value = "Unsuccessful"       
        vk.save(r'C:\Users\Chetan.K\Desktop\PTC_activity\Safa_Face_2\Code\Running_files\Safa_Primary1.xlsx')     
        driver.close()                
            
                        
        
    
                
                
               
                









