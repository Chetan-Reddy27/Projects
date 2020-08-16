from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl
import time
#storing state code
dic = {}
vk1 = openpyxl.load_workbook(r'C:\Users\Chetan.K\Desktop\PTC_activity\US_PTC_activity\State_code_match.xlsx')
sh1 = vk1.active
m = sh1.max_row
for i in range(2,m+1):
    dic[sh1.cell(i, 1).value] = sh1.cell(i, 2).value
    
vk = openpyxl.load_workbook(r'C:\Users\Chetan.K\Desktop\PTC_activity\Safa_Face_2\Code\Running_files\Safa_Secondary1.xlsx')
sh = vk.active
n = sh.max_row
print(n)

   
for i in range(2, n+1):
    try:
        driver = webdriver.Ie(r"C:\Users\Chetan.K\PycharmProjects\Office_Project\IEDriverServer.exe")
        value = sh.cell(i, 1).value
        driver.get(f'https://mdm.ey.net/businesspartnermaintenance/change/address/{value}')
        address_line1 = sh.cell(i, 5).value
        address_line2 = sh.cell(i, 6).value
        address_line3 = sh.cell(i, 7).value
        Country_hardcore = sh.cell(i, 8).value
        State = sh.cell(i, 9).value
        city = sh.cell(i, 10).value
        Postal_code = sh.cell(i, 11).value
        try:    
            driver.implicitly_wait(10)    
            driver.find_element_by_xpath("//*[@id='ddl-change-type-address']/option[2]").click()
            time.sleep(1)
            driver.find_element_by_xpath("//*[@id='Address_AddressLine1']").send_keys(address_line1)
            if not sh.cell(i, 6).value == None:
                driver.find_element_by_xpath("//*[@id='Address_AddressLine2']").send_keys(address_line2)
            if not sh.cell(i, 7).value == None:
                driver.find_element_by_xpath("//*[@id='Address_AddressLine3']").send_keys(address_line3)
            time.sleep(2)
            if Country_hardcore == 'CH':
                driver.find_element_by_id('ddl-country').send_keys('Switzerland')
            elif Country_hardcore == 'AT':
                driver.find_element_by_id('ddl-country').send_keys('Austria')
            else:
                driver.find_element_by_id('ddl-country').send_keys('Germany')
            time.sleep(2)
#            for k, v in dic.items():
#                if State == k:
#                    print(k, v)
#                    driver.find_element_by_xpath("//*[@id='ddl-state']").send_keys(v)
#                    time.sleep(1)
            if not sh.cell(i, 10).value == None:
                driver.find_element_by_xpath("//*[@id='Address_City']").send_keys(city)
            if not sh.cell(i, 11).value == None:
                driver.find_element_by_xpath("//*[@id='Address_PostalCode']").send_keys(Postal_code)
#            State_check = driver.find_element_by_id('ddl-state').get_attribute('value') 
            #driver.find_element_by_xpath("//*[@id='Address_PoBox']").send_keys(PO_box)
#            if not State_check == '':
            driver.find_element_by_id('btn-submit').send_keys(Keys.ENTER)
            time.sleep(2)
            print(i-1)
            sh.cell(i, 2).value = "successful"
            sh.cell(i, 4).value = "updated as secondary"
            vk.save(r'C:\Users\Chetan.K\Desktop\PTC_activity\Safa_Face_2\Code\Running_files\Safa_Secondary1.xlsx')
#            else:
#                print('state not updated')
#                sh.cell(i, 2).value = "unsuccessful"
#                sh.cell(i, 4).value = "state not updated"
#                vk.save(r'C:\Users\Chetan.K\Desktop\PTC_activity\Safa_Face_2\Code\Running_files\Safa_Secondary1.xlsx')
            
        except Exception as e:
            print('Some Issue with ' + sh.cell(i, 1).value)
            sh.cell(i, 2).value = "Unsuccessful"
            vk.save(r'C:\Users\Chetan.K\Desktop\PTC_activity\Safa_Face_2\Code\Running_files\Safa_Secondary1.xlsx')
            
        finally:
            driver.close()
    except Exception as e:
        print('some issue with internet')
        sh.cell(i, 2).value = "Unsuccessful"
        vk.save(r'C:\Users\Chetan.K\Desktop\PTC_activity\Safa_Face_2\Code\Running_files\Safa_Secondary1.xlsx')
#        driver.close()
        pass